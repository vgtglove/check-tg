import sys
import os
import asyncio
import time
import json  # 添加json模块导入
import configparser  # 添加到文件开头的import部分
from datetime import datetime, timedelta
from typing import List, Set, Dict, Tuple, Optional
import pandas as pd  # 用于Excel导出功能
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QPushButton, QTextEdit, QLabel,
                             QProgressBar, QFileDialog, QMessageBox, QLineEdit,
                             QFrame, QGroupBox, QSplitter, QTableWidget,
                             QTableWidgetItem, QHeaderView, QInputDialog, QMenu, QAction, QDialog, QGridLayout, QScrollArea, QCheckBox, QSizePolicy,
                             QMenuBar, QComboBox, QTabWidget, QRadioButton, QButtonGroup)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer, QDateTime, QPropertyAnimation, QEasingCurve, QRect, QPoint, QSize
from PyQt5.QtGui import QFont, QIcon, QPalette, QColor, QCursor, QBrush, QPainter, QPen, QRadialGradient
from telethon.sync import TelegramClient
from telethon.tl.functions.contacts import ImportContactsRequest
from telethon.tl.types import InputPhoneContact, UserStatusOnline, UserStatusOffline, UserStatusRecently, UserStatusLastWeek, UserStatusLastMonth
from telethon import functions, types
from ui.dialogs.dialogs import SendMessageDialog
from ui.ui_main import (
    PhoneNumberEdit, ConfigDialog, StyledProgressBar, AnimatedButton, StatusIndicator,
    AboutDialog, ContactDialog, ChangelogDialog, ActivityResultDialog, NumbersListWidget
)
from utils.phone_utils import normalize_phone_number, clean_phone_numbers
from utils.session_utils import SessionStatus
from utils.activity_utils import UserActivityStatus
from utils.threading_utils import CheckThread, ActivityCheckThread, PhoneNumberImportThread

# 启用高DPI缩放支持
os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

# 全局变量
client_id_counter = 0  # 用于生成唯一的联系人ID


class TelegramGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        # 加载配置
        self.config = configparser.ConfigParser()

        # 获取配置文件路径 - 优先使用可执行文件所在目录
        if getattr(sys, 'frozen', False):
            # 如果是打包后的exe运行
            base_dir = os.path.dirname(sys.executable)
            config_path = os.path.join(base_dir, 'config.ini')
        else:
            # 如果是开发环境运行
            config_path = 'config.ini'

        self.config.read(config_path, encoding='utf-8')

        self.sessions = []
        self.phone_numbers = []
        self.registered_numbers = set()
        self.settings_file = "telegram_settings.json"

        # 创建基本UI结构，但延迟加载重量级组件
        self.setupBasicUI()

        # 使用延迟加载机制
        QTimer.singleShot(100, self.finishInitialization)

    def setupBasicUI(self):
        """设置基本UI框架，快速显示窗口"""
        self.setWindowTitle('Telegram消息工具')
        self.setGeometry(100, 100, 1200, 800)

        # 应用全局样式表
        self.setStyleSheet("""
            QMainWindow {
                background-color: #F5F5F5;
            }
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                border: 1px solid #E0E0E0;
                border-radius: 6px;
                margin-top: 12px;
                padding-top: 10px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #424242;
                background-color: white;
            }
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                padding: 5px 15px;
                border-radius: 3px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:pressed {
                background-color: #0D47A1;
            }
            QPushButton:disabled {
                background-color: #BDBDBD;
            }
            QTextEdit, QLineEdit {
                border: 1px solid #E0E0E0;
                border-radius: 3px;
                padding: 5px;
                background-color: white;
            }
            QScrollBar:vertical {
                border: none;
                background: #F5F5F5;
                width: 8px;
                border-radius: 4px;
                margin: 0;
            }
            QScrollBar::handle:vertical {
                background-color: #BDBDBD;
                border-radius: 4px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #9E9E9E;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical,
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                height: 0;
                background: none;
            }
        """)

        # 显示"正在加载..."的标签
        loading_label = QLabel("正在加载组件和会话数据，请稍候...", self)
        loading_label.setAlignment(Qt.AlignCenter)
        loading_label.setStyleSheet("""
            font-size: 16px;
            color: #1976D2;
            background-color: #E3F2FD;
            border-radius: 8px;
            padding: 15px;
        """)
        loading_label.resize(400, 70)
        loading_label.move(int((self.width() - loading_label.width()) / 2),
                           int((self.height() - loading_label.height()) / 2))

        # 在标签下方添加进度条
        progress_bar = QProgressBar(self)
        progress_bar.setStyleSheet("""
            QProgressBar {
                border: none;
                border-radius: 5px;
                text-align: center;
                height: 20px;
                background-color: #F5F5F5;
            }
            QProgressBar::chunk {
                background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                                stop:0 #2196F3, stop:1 #4CAF50);
                border-radius: 5px;
            }
        """)
        progress_bar.resize(400, 20)
        progress_bar.move(int((self.width() - progress_bar.width()) / 2),
                          int((self.height() - loading_label.height()) / 2) + 80)

        # 设置属性以便后续引用
        self.loading_label = loading_label
        self.loading_progress = progress_bar

        # 设置进度条动画
        self.loading_timer = QTimer(self)
        self.loading_timer.timeout.connect(self.updateLoadingProgress)
        self.loading_progress.setValue(0)
        self.loading_timer.start(50)  # 每50毫秒更新一次

    def updateLoadingProgress(self):
        """更新加载进度条"""
        current_value = self.loading_progress.value()
        if current_value < 100:
            self.loading_progress.setValue(current_value + 1)
        else:
            self.loading_timer.stop()

    def finishInitialization(self):
        """完成剩余的初始化工作"""
        try:
            # 加载设置和已注册号码
            self.load_settings()
            self.load_registered_numbers()

            # 初始化完整UI
            self.initUI()

            # 隐藏加载界面
            if hasattr(self, 'loading_label') and self.loading_label:
                self.loading_label.hide()
            if hasattr(self, 'loading_progress') and self.loading_progress:
                self.loading_progress.hide()

            # 完成进度条动画
            if hasattr(self, 'loading_timer') and self.loading_timer and self.loading_timer.isActive():
                self.loading_timer.stop()

            # 在状态栏显示初始化完成消息
            try:
                self.statusBar().showMessage("初始化完成!", 3000)  # 显示3秒
            except:
                pass

            # 延迟预加载表格数据，减轻主线程负担
            QTimer.singleShot(500, self.delayed_load_data)

        except Exception as e:
            import traceback
            error_msg = f"完成初始化时出错: {str(e)}\n{traceback.format_exc()}"
            print(error_msg)
            try:
                QMessageBox.critical(self, "初始化错误", f"完成初始化时出错:\n\n{str(e)}")
            except:
                pass

    def delayed_load_data(self):
        """延迟加载数据"""
        try:
            # 预加载表格数据
            self.update_session_list()
        except Exception as e:
            import traceback
            error_msg = f"加载数据时出错: {str(e)}\n{traceback.format_exc()}"
            try:
                self.log(f"【加载错误】{error_msg}")
            except:
                print(error_msg)

    def initUI(self):
        # 创建菜单栏
        self.createMenuBar()

        # 主窗口部件和布局
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)

        # ======== 创建上中下三区域分割的主布局 ========
        main_splitter = QSplitter(Qt.Vertical)

        # ======== 顶部区域 - Session管理 ========
        top_panel = QWidget()
        top_layout = QVBoxLayout(top_panel)
        top_layout.setContentsMargins(0, 0, 0, 0)

        # Session管理组
        session_group = QGroupBox(f'Session管理 (共 {len(self.sessions)} 个)')
        session_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                border: 2px solid #E0E0E0;
                border-radius: 6px;
                margin-top: 12px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #424242;
            }
        """)
        session_layout = QVBoxLayout()
        session_layout.setSpacing(10)

        # 工具栏区域
        toolbar = QWidget()
        toolbar.setStyleSheet("""
            QWidget {
                background-color: #F5F5F5;
                border-radius: 6px;
            }
        """)
        toolbar_layout = QHBoxLayout(toolbar)
        toolbar_layout.setContentsMargins(10, 5, 10, 5)
        toolbar_layout.setSpacing(15)

        # 创建工具栏按钮
        def create_toolbar_button(icon, text, tooltip, color):
            btn = AnimatedButton(text, color)
            # 删除图标设置代码
            btn.setToolTip(tooltip)

            # 确保按钮有足够空间显示文本
            font_metrics = btn.fontMetrics()
            text_width = font_metrics.width(f" {text}")
            btn.setMinimumWidth(max(120, text_width + 40))  # 文本宽度加上边距的空间

            return btn

        # 全选按钮
        select_all_btn = create_toolbar_button(
            "select-all", "全选/取消",
            "全选或取消选择所有会话",
            "#673AB7"
        )
        select_all_btn.clicked.connect(self.toggle_select_all_sessions)

        # 扫描Session按钮
        scan_btn = create_toolbar_button(
            "scan", "扫描Session",
            "扫描并添加sessions目录下的所有.session文件\n这些文件是您的Telegram登录凭证",
            "#2196F3"
        )
        scan_btn.clicked.connect(self.add_session)

        # 移除Session按钮
        remove_btn = create_toolbar_button(
            "remove", "移除Session",
            "从列表中移除选中的会话文件\n不会删除实际的.session文件",
            "#FF5722"
        )
        remove_btn.clicked.connect(self.remove_session)

        # 批量移除按钮
        batch_remove_btn = create_toolbar_button(
            "batch-remove", "批量移除",
            "同时移除多个选中的会话文件\n可以批量清理不需要的会话",
            "#E91E63"
        )
        batch_remove_btn.clicked.connect(self.batch_remove_sessions)

        # 删除不可用session按钮
        delete_invalid_btn = create_toolbar_button(
            "delete", "删除不可用",
            "删除所有错误或未授权的会话\n同时删除对应的.session文件",
            "#F44336"
        )
        delete_invalid_btn.clicked.connect(self.delete_invalid_sessions)

        # 添加按钮到工具栏
        toolbar_layout.addWidget(select_all_btn)
        toolbar_layout.addWidget(scan_btn)
        toolbar_layout.addWidget(remove_btn)
        toolbar_layout.addWidget(batch_remove_btn)
        toolbar_layout.addWidget(delete_invalid_btn)
        toolbar_layout.addStretch()

        # Session列表区域 - 使用QScrollArea和QTableWidget
        self.session_table = QTableWidget()
        self.session_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #E0E0E0;
                background-color: white;
                gridline-color: #F0F0F0;
            }
            QTableWidget::item {
                padding: 4px;
                border-bottom: 1px solid #F0F0F0;
            }
            QTableWidget::item:selected {
                background-color: #E3F2FD;
                color: #000000;
            }
            QHeaderView::section {
                background-color: #F5F5F5;
                padding: 6px;
                border: none;
                border-right: 1px solid #E0E0E0;
                border-bottom: 1px solid #E0E0E0;
                font-weight: bold;
                color: #424242;
            }
        """)

        # 设置列
        headers = ["选择", "会话名称", "状态", "冷却时间", "批量大小", "检测次数", "错误次数", "操作"]
        self.session_table.setColumnCount(len(headers))
        self.session_table.setHorizontalHeaderLabels(headers)

        # 设置表格属性
        self.session_table.setSelectionBehavior(
            QTableWidget.SelectRows)  # 整行选择
        self.session_table.setEditTriggers(
            QTableWidget.NoEditTriggers)   # 不可编辑
        self.session_table.horizontalHeader().setHighlightSections(False)  # 取消表头高亮
        self.session_table.verticalHeader().setVisible(False)  # 隐藏行号

        # 设置列宽模式
        self.session_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.Fixed)  # 选择列固定宽度
        self.session_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.Stretch)  # 名称列自适应
        self.session_table.setColumnWidth(0, 50)  # 设置选择列宽度

        # 将表格添加到滚动区域
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: transparent;
            }
            QScrollBar:vertical {
                border: none;
                background: #F5F5F5;
                width: 8px;
                border-radius: 4px;
            }
            QScrollBar::handle:vertical {
                background: #BDBDBD;
                border-radius: 4px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background: #9E9E9E;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }
        """)

        # 创建容器以包含表格，这是个重要的性能优化，避免直接在滚动区域中重绘大表格
        table_container = QWidget()
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(0, 0, 0, 0)
        table_layout.addWidget(self.session_table)

        scroll_area.setWidget(table_container)

        # 组装Session管理区域
        session_layout.addWidget(toolbar)
        session_layout.addWidget(scroll_area)
        session_group.setLayout(session_layout)
        top_layout.addWidget(session_group)

        # 更新会话列表
        self.update_session_list()

        # ======== 中部区域 - 号码输入和显示 ========
        middle_panel = QWidget()
        middle_layout = QHBoxLayout(middle_panel)
        middle_layout.setContentsMargins(0, 0, 0, 0)
        middle_layout.setSpacing(10)

        # 左侧 - 号码输入区域
        input_group = QGroupBox('号码输入')
        input_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                border: 2px solid #E0E0E0;
                color: #000000;
                border-radius: 6px;
                margin-top: 12px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #424242;
            }
        """)

        input_layout = QVBoxLayout()
        input_layout.setSpacing(10)

        # 输入框和按钮的布局
        input_container = QHBoxLayout()
        input_container.setSpacing(15)

        # 左侧输入区域
        input_left = QVBoxLayout()
        self.phone_input = PhoneNumberEdit()
        self.phone_input.setStyleSheet("""
            QTextEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 8px;
                font-size: 13px;
                background-color: white;
            }
            QTextEdit:focus {
                border: 1px solid #2196F3;
            }
        """)
        self.phone_input.setPlaceholderText(
            '在此输入手机号码，每行一个\n支持格式：国际格式(+86...)或国内格式(139...)')
        self.phone_input.setMinimumHeight(150)

        # 格式说明标签
        format_label = QLabel("支持格式: 国际格式(+86...)、国内格式(139...)，自动格式化")
        format_label.setStyleSheet("""
            QLabel {
                color: #757575;
                font-style: italic;
                font-size: 12px;
                padding: 5px;
            }
        """)

        input_left.addWidget(self.phone_input)
        input_left.addWidget(format_label)

        # 右侧按钮区域
        input_right = QVBoxLayout()
        input_right.setSpacing(10)

        def create_action_button(icon, text, tooltip, color):
            btn = AnimatedButton(text, color)
            # 删除图标设置代码
            btn.setToolTip(tooltip)

            # 确保按钮有足够空间显示文本
            font_metrics = btn.fontMetrics()
            text_width = font_metrics.width(f" {text}")
            btn.setMinimumWidth(max(120, text_width + 40))  # 文本宽度加上边距的空间

            return btn

        # 添加号码按钮
        add_number_btn = create_action_button(
            "add", "添加号码",
            "将输入框中的号码添加到检测列表\n会自动过滤重复和已注册的号码",
            "#4CAF50"
        )
        add_number_btn.clicked.connect(self.add_phone_number)

        # 从文件导入按钮
        import_btn = create_action_button(
            "import", "从文件导入",
            "从文本文件导入号码\n文件中每行一个号码\n支持txt格式文件",
            "#2196F3"
        )
        import_btn.clicked.connect(self.import_numbers)

        # 清空号码按钮
        clear_numbers_btn = create_action_button(
            "clear", "清空号码",
            "清空当前已添加的所有待检测号码",
            "#FF5722"
        )
        clear_numbers_btn.clicked.connect(self.clear_phone_numbers)

        input_right.addWidget(add_number_btn)
        input_right.addWidget(import_btn)
        input_right.addWidget(clear_numbers_btn)
        input_right.addStretch()

        # 组装输入区域
        input_container.addLayout(input_left, 7)
        input_container.addLayout(input_right, 3)
        input_layout.addLayout(input_container)

        # 提示信息区域
        note_container = QWidget()
        note_container.setStyleSheet("""
            QWidget {
                background-color: #E3F2FD;
                border-radius: 4px;
                padding: 5px;
            }
        """)
        note_layout = QHBoxLayout(note_container)
        note_layout.setContentsMargins(10, 5, 10, 5)

        note_icon = QLabel("ℹ")
        note_icon.setStyleSheet("""
            QLabel {
                color: #1976D2;
                font-size: 16px;
                font-weight: bold;
            }
        """)

        note_text = QLabel("重复和已注册号码会被自动过滤，不会添加到检测列表")
        note_text.setStyleSheet("""
            QLabel {
                color: #1976D2;
                font-style: italic;
                font-size: 12px;
            }
        """)

        note_layout.addWidget(note_icon)
        note_layout.addWidget(note_text, 1)
        note_layout.addStretch()

        input_layout.addWidget(note_container)
        input_group.setLayout(input_layout)

        # 右侧 - 已添加号码显示
        numbers_group = QGroupBox('已添加的号码')
        numbers_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                border: 2px solid #E0E0E0;
                border-radius: 6px;
                margin-top: 12px;
                color: #424242;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #424242;
            }
        """)

        numbers_layout = QVBoxLayout()

        # 使用优化的号码列表显示组件
        self.numbers_list = NumbersListWidget()
        self.numbers_list.setStyleSheet("""
            QTextEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 8px;
                font-size: 13px;
                background-color: #FAFAFA;
            }
        """)

        numbers_layout.addWidget(self.numbers_list)
        numbers_group.setLayout(numbers_layout)

        # 组装中部区域
        middle_layout.addWidget(input_group, 5)
        middle_layout.addWidget(numbers_group, 5)

        # ======== 底部区域 - 运行状态和控制 ========
        bottom_panel = QWidget()
        bottom_layout = QVBoxLayout(bottom_panel)
        bottom_layout.setContentsMargins(0, 0, 0, 0)

        # 状态显示组
        status_group = QGroupBox('运行状态')
        status_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                border: 2px solid #E0E0E0;
                border-radius: 6px;
                margin-top: 12px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #424242;
            }
        """)

        status_layout = QVBoxLayout()

        # 状态文本显示
        self.status_text = QTextEdit()
        self.status_text.setReadOnly(True)
        self.status_text.setStyleSheet("""
            QTextEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 8px;
                font-size: 13px;
                background-color: #FAFAFA;
                font-family: Consolas, Monaco, monospace;
            }
        """)

        # 控制区域
        control_panel = QWidget()
        control_panel.setStyleSheet("""
            QWidget {
                background-color: #F5F5F5;
                border-radius: 6px;
            }
        """)
        control_layout = QVBoxLayout(control_panel)  # 改为垂直布局
        control_layout.setContentsMargins(10, 5, 10, 5)

        # 进度条区域
        progress_container = QWidget()
        progress_layout = QVBoxLayout(progress_container)
        progress_layout.setContentsMargins(0, 0, 0, 0)

        progress_label = QLabel("检测进度:")
        progress_label.setStyleSheet("""
            QLabel {
                color: #424242;
                font-size: 12px;
                font-weight: bold;
            }
        """)

        self.progress_bar = StyledProgressBar()

        progress_layout.addWidget(progress_label)
        progress_layout.addWidget(self.progress_bar)

        # 控制按钮区域
        buttons_container = QWidget()
        buttons_layout = QHBoxLayout(buttons_container)
        buttons_layout.setSpacing(10)

        def create_control_button(icon, text, tooltip, color):
            btn = AnimatedButton(text, color)
            # 删除图标设置代码
            btn.setToolTip(tooltip)
            return btn

        # 开始检查按钮
        self.start_btn = create_control_button(
            "start", "开始检查",
            "开始检测已添加的号码\n请确保有可用的会话且已添加号码",
            "#4CAF50"
        )
        self.start_btn.clicked.connect(self.start_check)

        # 检测活跃度按钮
        self.check_activity_btn = create_control_button(
            "activity", "检测活跃度",
            "检测已添加号码的活跃度\n可以了解用户最近是否在线",
            "#9C27B0"
        )
        self.check_activity_btn.clicked.connect(self.start_activity_check)

        # 停止按钮
        self.stop_btn = create_control_button(
            "stop", "停止",
            "停止当前的检测过程\n会在完成当前批次后停止",
            "#FF5722"
        )
        self.stop_btn.setEnabled(False)
        self.stop_btn.clicked.connect(self.stop_check)

        # 打开结果位置按钮
        self.open_result_location_btn = create_control_button(
            "folder", "打开结果位置",
            "打开保存结果的文件夹\n可以查看已注册号码的记录",
            "#607D8B"
        )
        self.open_result_location_btn.clicked.connect(
            self.open_result_file_location)

        # 配置按钮
        self.config_btn = create_control_button(
            "settings", "配置",
            "打开配置窗口\n可以设置API参数和默认值",
            "#9C27B0"
        )
        self.config_btn.clicked.connect(self.show_config_dialog)

        buttons_layout.addWidget(self.start_btn)
        buttons_layout.addWidget(self.check_activity_btn)
        buttons_layout.addWidget(self.stop_btn)
        buttons_layout.addWidget(self.open_result_location_btn)
        buttons_layout.addWidget(self.config_btn)
        buttons_layout.addStretch()

        # 组装控制面板 - 使用垂直布局，进度条在上，按钮在下
        control_layout.addWidget(progress_container)
        control_layout.addWidget(buttons_container)

        # 组装状态区域
        status_layout.addWidget(self.status_text)
        status_layout.addWidget(control_panel)
        status_group.setLayout(status_layout)
        bottom_layout.addWidget(status_group)

        # 将三个主要区域添加到主分割器
        main_splitter.addWidget(top_panel)
        main_splitter.addWidget(middle_panel)
        main_splitter.addWidget(bottom_panel)

        # 设置初始分割比例
        main_splitter.setSizes([250, 250, 300])

        # 添加主分割器到主布局
        layout.addWidget(main_splitter)
        main_widget.setLayout(layout)

        # 添加定时器更新session状态
        self.status_timer = QTimer()
        self.status_timer.timeout.connect(self.update_session_list)
        self.status_timer.start(1000)  # 每秒更新一次

        self.worker = None

    def createMenuBar(self):
        """创建菜单栏"""
        menubar = self.menuBar()
        menubar.setStyleSheet("""
            QMenuBar {
                background-color: #F5F5F5;
                border-bottom: 1px solid #E0E0E0;
                font-size: 13px;
            }
            QMenuBar::item {
                padding: 5px 10px;
                background: transparent;
            }
            QMenuBar::item:selected {
                background: #E3F2FD;
                border-radius: 4px;
            }
            QMenuBar::item:pressed {
                background: #BBDEFB;
                border-radius: 4px;
            }
            QMenu {
                background-color: white;
                border: 1px solid #E0E0E0;
                border-radius: 4px;
            }
            QMenu::item {
                padding: 5px 25px 5px 20px;
                border-radius: 4px;
            }
            QMenu::item:selected {
                background-color: #E3F2FD;
                color: #1976D2;
            }
        """)

        # 关于菜单项
        about_action = QAction('关于', self)
        about_action.triggered.connect(self.showAboutDialog)
        menubar.addAction(about_action)

        # 联系我们菜单项
        contact_action = QAction('联系我们', self)
        contact_action.triggered.connect(self.showContactDialog)
        menubar.addAction(contact_action)

        # 更新日志菜单项
        changelog_action = QAction('更新日志', self)
        changelog_action.triggered.connect(self.showChangelogDialog)
        menubar.addAction(changelog_action)

    def showAboutDialog(self):
        """显示关于对话框"""
        dialog = AboutDialog(self)
        dialog.exec_()

    def showContactDialog(self):
        """显示联系我们对话框"""
        dialog = ContactDialog(self)
        dialog.exec_()

    def showChangelogDialog(self):
        """显示更新日志对话框"""
        dialog = ChangelogDialog(self)
        dialog.exec_()

    def add_session(self):
        """扫描并添加新的session文件"""
        # 优先使用环境变量中设置的sessions目录路径
        if 'SESSIONS_DIR' in os.environ and os.environ['SESSIONS_DIR']:
            base_dir = os.environ['SESSIONS_DIR']
        else:
            # 如果环境变量未设置，则使用相对路径
            base_dir = os.path.join(os.path.dirname(
                os.path.abspath(__file__)), 'sessions')

        # 如果sessions目录不存在则创建
        if not os.path.exists(base_dir):
            os.makedirs(base_dir)
            QMessageBox.information(self, '提示', f'已创建sessions目录: {base_dir}')
            return

        # 递归搜索所有.session文件
        session_files = []
        for root, dirs, files in os.walk(base_dir):
            for file in files:
                if file.endswith('.session'):
                    full_path = os.path.join(root, file)
                    # 检查是否已添加
                    if not any(s.file_path == full_path for s in self.sessions):
                        session_files.append(full_path)

        if not session_files:
            QMessageBox.warning(
                self, '警告', f'sessions目录下未找到任何.session文件！\n路径: {base_dir}')
            return

        # 添加找到的所有新session文件，使用配置文件中的默认值
        added_count = 0
        for file_path in session_files:
            # 创建新的SessionStatus对象时使用配置文件的默认值
            self.sessions.append(SessionStatus(
                file_path, use_defaults=True, config=self.config))
            added_count += 1

        self.update_session_list()
        if added_count > 0:
            QMessageBox.information(
                self, '成功', f'已添加 {added_count} 个新的Session文件！')
            self.save_settings()  # 保存设置

    def remove_session(self):
        """移除选中的Session"""
        if not self.sessions:
            QMessageBox.warning(self, '警告', '没有可移除的Session！')
            return

        # 显示Session列表供用户选择
        items = [session.name for session in self.sessions]
        item, ok = QInputDialog.getItem(self, "选择Session",
                                        "请选择要移除的Session：",
                                        items, 0, False)
        if ok and item:
            # 找到对应的session
            for session in self.sessions:
                if session.name == item:
                    reply = QMessageBox.question(self, '确认',
                                                 f'确定要移除Session {session.name} 吗？',
                                                 QMessageBox.Yes | QMessageBox.No)
                    if reply == QMessageBox.Yes:
                        self.sessions.remove(session)
                        self.update_session_list()
                        self.log(f"已移除Session: {session.name}")
                        self.save_settings()
                    break

    def batch_remove_sessions(self):
        """批量移除Session"""
        if not self.sessions:
            QMessageBox.warning(self, '警告', '没有可移除的Session！')
            return

        # 获取所有被选中的session
        selected_sessions = []
        for session in self.sessions:
            if session.file_path in self.selected_sessions and self.selected_sessions[session.file_path].isChecked():
                selected_sessions.append(session)

        if not selected_sessions:
            QMessageBox.warning(self, '警告', '未选择任何Session！请在表格中勾选要删除的Session。')
            return

        reply = QMessageBox.question(
            self,
            '确认',
            f'确定要移除选中的 {len(selected_sessions)} 个Session吗？',
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            for session in selected_sessions:
                self.sessions.remove(session)

            self.update_session_list()
            self.log(f"已批量移除 {len(selected_sessions)} 个Session")
            self.save_settings()
            QMessageBox.information(self, '成功',
                                    f'已成功移除 {len(selected_sessions)} 个Session')

    def update_session_list(self):
        """更新会话列表显示"""
        try:
            # 暂停定时器以防止更新过程中的干扰
            if hasattr(self, 'status_timer') and self.status_timer and self.status_timer.isActive():
                self.status_timer.stop()

            # 获取当前垂直滚动条位置
            scroll_value = 0
            if hasattr(self, 'session_table') and self.session_table:
                scrollbar = self.session_table.verticalScrollBar()
                if scrollbar:
                    scroll_value = scrollbar.value()

            # 保存当前勾选状态
            checked_sessions = set()
            if hasattr(self, 'selected_sessions'):
                for file_path, checkbox in self.selected_sessions.items():
                    if checkbox and checkbox.isChecked():
                        checked_sessions.add(file_path)

            # 更新会话管理组标题
            for widget in self.findChildren(QGroupBox):
                if widget and widget.title() and widget.title().startswith('Session管理'):
                    widget.setTitle(f'Session管理 (共 {len(self.sessions)} 个)')
                    break

            # 确保session_table已创建
            if not hasattr(self, 'session_table') or not self.session_table:
                return

            # 设置表格行数
            try:
                self.session_table.setRowCount(len(self.sessions))
            except Exception as e:
                print(f"设置表格行数时出错: {str(e)}")
                return

            # 设置行高
            try:
                self.session_table.verticalHeader().setDefaultSectionSize(40)
            except Exception as e:
                print(f"设置行高时出错: {str(e)}")

            # 创建勾选状态字典
            self.selected_sessions = {}

            # 阻塞表格信号，减少重绘次数
            try:
                self.session_table.blockSignals(True)
                self.session_table.setUpdatesEnabled(False)
            except Exception as e:
                print(f"阻塞表格信号时出错: {str(e)}")

            # 计算可见行的范围 - 只更新当前可见的行和周围的一些行
            visible_rows = min(20, len(self.sessions))  # 默认显示前20行或全部(如果少于20个)

            # 只更新可见范围内的行，其他行延迟加载
            visible_start = max(0, scroll_value // 40)  # 40是行高
            visible_end = min(len(self.sessions), visible_start + visible_rows)

            # 更新会话表格中的行 - 只处理可见行
            for row in range(visible_start, visible_end):
                try:
                    self.update_session_row(row, checked_sessions)
                except Exception as row_error:
                    print(f"更新行 {row} 时出错: {str(row_error)}")

            # 恢复表格信号和更新
            try:
                self.session_table.setUpdatesEnabled(True)
                self.session_table.blockSignals(False)
            except Exception as e:
                print(f"恢复表格信号时出错: {str(e)}")

            # 一次性重绘表格，而不是每个单元格都触发重绘
            try:
                self.session_table.viewport().update()
            except Exception as e:
                print(f"更新表格视图时出错: {str(e)}")

            # 恢复滚动条位置
            try:
                QTimer.singleShot(
                    10, lambda: self.restore_scroll_position(scroll_value))
            except Exception as e:
                print(f"恢复滚动条位置时出错: {str(e)}")

            # 如果有很多session，延迟加载剩余部分
            if len(self.sessions) > visible_rows:
                try:
                    QTimer.singleShot(100, lambda: self.load_remaining_rows(
                        visible_start, visible_end, checked_sessions))
                except Exception as e:
                    print(f"延迟加载剩余行时出错: {str(e)}")
            else:
                # 重新启动定时器，但使用更长的间隔
                try:
                    QTimer.singleShot(100, lambda: self.start_status_timer())
                except Exception as e:
                    print(f"启动状态定时器时出错: {str(e)}")

        except Exception as e:
            import traceback
            error_msg = f"更新会话列表时出错: {str(e)}\n{traceback.format_exc()}"
            print(error_msg)
            try:
                if hasattr(self, 'log'):
                    self.log(f"【更新错误】{error_msg}")
            except:
                pass

            # 无论如何都尝试重启状态定时器
            try:
                QTimer.singleShot(2000, lambda: self.start_status_timer())
            except:
                pass

    def load_remaining_rows(self, visible_start, visible_end, checked_sessions):
        """延迟加载剩余行"""
        try:
            if not hasattr(self, 'session_table') or not self.session_table:
                return

            # 阻塞信号
            try:
                self.session_table.blockSignals(True)
                self.session_table.setUpdatesEnabled(False)
            except Exception as e:
                print(f"阻塞信号时出错: {str(e)}")
                return

            # 处理不可见区域的行 - 每批次处理50个，避免UI阻塞
            batch_size = 50
            total_rows = len(self.sessions)

            try:
                # 分批处理前面部分
                for row in range(0, min(visible_start, total_rows), batch_size):
                    if not hasattr(self, 'session_table') or not self.session_table:
                        break
                    end_row = min(row + batch_size, visible_start)
                    for i in range(row, end_row):
                        try:
                            self.update_session_row(i, checked_sessions)
                        except Exception as row_error:
                            print(f"更新行 {i} 时出错: {str(row_error)}")
                    # 给UI一个短暂的响应时间
                    QApplication.processEvents()

                # 分批处理后面部分
                for row in range(visible_end, total_rows, batch_size):
                    if not hasattr(self, 'session_table') or not self.session_table:
                        break
                    end_row = min(row + batch_size, total_rows)
                    for i in range(row, end_row):
                        try:
                            self.update_session_row(i, checked_sessions)
                        except Exception as row_error:
                            print(f"更新行 {i} 时出错: {str(row_error)}")
                    # 给UI一个短暂的响应时间
                    QApplication.processEvents()
            except Exception as batch_error:
                print(f"批量处理行时出错: {str(batch_error)}")

            # 恢复信号
            try:
                self.session_table.setUpdatesEnabled(True)
                self.session_table.blockSignals(False)
            except Exception as e:
                print(f"恢复信号时出错: {str(e)}")

            # 重启定时器
            try:
                self.start_status_timer()
            except Exception as e:
                print(f"重启状态定时器时出错: {str(e)}")

        except Exception as e:
            import traceback
            error_msg = f"加载剩余行时出错: {str(e)}\n{traceback.format_exc()}"
            print(error_msg)
            # 无论如何都尝试重启状态定时器
            try:
                QTimer.singleShot(2000, self.start_status_timer)
            except:
                pass

    def update_session_row(self, row, checked_sessions):
        """更新单个会话行"""
        if row >= len(self.sessions):
            return

        session = self.sessions[row]

        # 选择列 - 添加勾选框
        checkbox = QCheckBox()
        checkbox.setStyleSheet("""
            QCheckBox {
                margin-left: 10px;
            }
            QCheckBox::indicator {
                width: 16px;
                height: 16px;
            }
            QCheckBox::indicator:unchecked {
                border: 2px solid #E0E0E0;
                background-color: white;
                border-radius: 3px;
            }
            QCheckBox::indicator:checked {
                border: 2px solid #2196F3;
                background-color: #2196F3;
                border-radius: 3px;
            }
            QCheckBox::indicator:hover {
                border: 2px solid #2196F3;
            }
        """)
        # 恢复之前的勾选状态
        if session.file_path in checked_sessions:
            checkbox.setChecked(True)

        checkbox_widget = QWidget()
        checkbox_layout = QHBoxLayout(checkbox_widget)
        checkbox_layout.addWidget(checkbox)
        checkbox_layout.setAlignment(Qt.AlignCenter)
        checkbox_layout.setContentsMargins(0, 0, 0, 0)
        self.session_table.setCellWidget(row, 0, checkbox_widget)

        # 保存checkbox引用
        self.selected_sessions[session.file_path] = checkbox

        # 会话名称
        name_item = QTableWidgetItem(session.name)
        name_item.setToolTip(f'会话文件路径: {session.file_path}')
        self.session_table.setItem(row, 1, name_item)

        # 状态
        status_colors = {
            "错误": "#FF5722",
            "未授权": "#FF5722",
            "正在运行": "#4CAF50",
            "空闲": "#2196F3"
        }

        status_text = ""
        if session.status == "错误":
            status_text = "错误"
        elif session.status == "未授权":
            status_text = "未授权"
        elif session.status == "正在运行":
            status_text = "正在运行"
        else:
            status_text = "空闲"

        # 使用自定义状态指示器
        status_widget = QWidget()
        status_layout = QHBoxLayout(status_widget)
        status_layout.setContentsMargins(5, 0, 5, 0)
        status_layout.setSpacing(8)

        # 创建状态指示器
        indicator = StatusIndicator(status_text)

        # 状态文字
        status_label = QLabel(status_text)
        status_label.setStyleSheet(
            f"color: {status_colors.get(status_text, '#757575')}; font-weight: bold;")

        status_layout.addWidget(indicator)
        status_layout.addWidget(status_label)
        status_layout.addStretch()

        self.session_table.setCellWidget(row, 2, status_widget)

        # 冷却时间
        cooldown_item = QTableWidgetItem(f"{session.cooldown_time}秒")
        cooldown_item.setTextAlignment(Qt.AlignCenter)
        self.session_table.setItem(row, 3, cooldown_item)

        # 批量大小
        batch_item = QTableWidgetItem(str(session.batch_size))
        batch_item.setTextAlignment(Qt.AlignCenter)
        self.session_table.setItem(row, 4, batch_item)

        # 检测次数
        checks_item = QTableWidgetItem(str(session.total_checks))
        checks_item.setTextAlignment(Qt.AlignCenter)
        self.session_table.setItem(row, 5, checks_item)

        # 错误次数
        error_item = QTableWidgetItem(str(session.error_count))
        error_item.setTextAlignment(Qt.AlignCenter)
        if session.error_count > 0:
            error_item.setForeground(QColor("#FF5722"))
        self.session_table.setItem(row, 6, error_item)

        # 操作按钮 - 使用动画按钮
        settings_btn = AnimatedButton("设置", "#2196F3")
        settings_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 3px;
                padding: 4px 10px;
                font-size: 12px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:pressed {
                background-color: #0D47A1;
            }
        """)

        # 使用lambda函数传递session对象而不是索引
        settings_btn.clicked.connect(
            lambda checked=False, s=session: self.show_session_settings_menu(s))
        settings_btn.setToolTip("设置会话参数")

        btn_widget = QWidget()
        btn_layout = QHBoxLayout(btn_widget)
        btn_layout.setContentsMargins(2, 2, 2, 2)
        btn_layout.addWidget(settings_btn)
        btn_layout.setAlignment(Qt.AlignCenter)

        self.session_table.setCellWidget(row, 7, btn_widget)

        # 添加行交替颜色
        if row % 2 == 0:
            for col in range(self.session_table.columnCount()):
                if self.session_table.item(row, col):
                    self.session_table.item(
                        row, col).setBackground(QColor("#F9F9F9"))

    def start_status_timer(self):
        """以更低的频率启动状态更新定时器"""
        if hasattr(self, 'status_timer'):
            self.status_timer.start(3000)  # 降低更新频率，改为3秒一次

    def restore_scroll_position(self, position):
        """恢复滚动条位置"""
        if hasattr(self, 'session_table') and self.session_table:
            scrollbar = self.session_table.verticalScrollBar()
            if scrollbar:
                scrollbar.setValue(position)

    def show_session_settings_menu(self, session):
        """显示会话设置菜单"""
        menu = QMenu(self)

        # 设置冷却时间
        cooldown_action = QAction("设置冷却时间", self)
        cooldown_action.triggered.connect(
            lambda: self.set_session_cooldown(self.sessions.index(session)))
        menu.addAction(cooldown_action)

        # 设置批量大小
        batch_action = QAction("设置批量大小", self)
        batch_action.triggered.connect(
            lambda: self.set_session_batch_size(self.sessions.index(session)))
        menu.addAction(batch_action)

        menu.addSeparator()

        # 移除会话
        remove_action = QAction("移除Session", self)
        remove_action.triggered.connect(
            lambda: self.remove_specific_session(session))
        menu.addAction(remove_action)

        # 在鼠标位置显示菜单
        cursor_pos = QCursor.pos()
        menu.exec_(cursor_pos)

    def remove_specific_session(self, session):
        """移除特定的Session"""
        reply = QMessageBox.question(
            self,
            '确认',
            f'确定要移除Session {session.name} 吗？',
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            self.sessions.remove(session)
            self.update_session_list()
            self.log(f"已移除Session: {session.name}")
            self.save_settings()

    def set_session_cooldown(self, row):
        """设置会话冷却时间"""
        session = self.sessions[row]
        current_cooldown = session.cooldown_time

        cooldown, ok = QInputDialog.getInt(
            self, "设置冷却时间",
            "请输入冷却时间（秒）：",
            current_cooldown, 1, 3600, 1
        )

        if ok:
            session.cooldown_time = cooldown
            self.update_session_list()
            self.log(f"已将 {session.name} 的冷却时间设置为 {cooldown} 秒")
            self.save_settings()  # 保存设置

    def set_session_batch_size(self, row):
        """设置会话批量大小"""
        session = self.sessions[row]
        current_batch = session.batch_size

        batch_size, ok = QInputDialog.getInt(
            self, "设置批量大小",
            "请输入每次检测的号码数量：",
            current_batch, 1, 100, 1
        )

        if ok:
            session.batch_size = batch_size
            self.update_session_list()
            self.log(f"已将 {session.name} 的批量大小设置为 {batch_size}")
            self.save_settings()  # 保存设置

    def load_registered_numbers(self):
        """加载已注册的号码,并统一格式化处理"""
        # 获取phone_register.txt文件路径 - 优先使用可执行文件所在目录
        if getattr(sys, 'frozen', False):
            # 如果是打包后的exe运行
            base_dir = os.path.dirname(sys.executable)
            phone_register_path = os.path.join(base_dir, 'phone_register.txt')
        else:
            # 如果是开发环境运行
            phone_register_path = 'phone_register.txt'

        try:
            with open(phone_register_path, 'r') as f:
                # 统一格式化处理每个号码
                self.registered_numbers = {self.normalize_phone_number(line.strip())
                                           for line in f if line.strip()}
        except FileNotFoundError:
            self.registered_numbers = set()
            # 创建文件
            try:
                with open(phone_register_path, 'w') as f:
                    f.write("")  # 创建一个空文件
            except Exception as e:
                self.log(f"无法创建phone_register.txt: {str(e)}")
                pass

    def normalize_phone_number(self, number):
        """统一号码格式,移除所有非数字字符"""
        # 提取所有数字
        digits = ''.join(filter(str.isdigit, number))

        # 如果第一位是1且总长度大于10,认为是带国家代码的号码
        if digits.startswith('1') and len(digits) > 10:
            return '+' + digits
        # 如果长度大于10但不以1开头,可能是其他国家号码
        elif len(digits) > 10:
            return '+' + digits
        # 如果只有10位,默认添加+1
        elif len(digits) == 10:
            return '+1' + digits
        else:
            return '+' + digits

    def add_phone_number(self):
        text = self.phone_input.toPlainText().strip()
        if not text:
            QMessageBox.warning(self, '警告', '请输入手机号码！')
            return

        # 分割输入的文本（按行分割）
        numbers = []
        already_registered = []
        for line in text.split('\n'):
            # 预处理：移除括号和多余空格
            line = line.replace('(', '').replace(')', '').strip()
            if not line:
                continue

            # 统一格式化号码
            normalized_number = self.normalize_phone_number(line)

            # 检查是否已在注册列表中
            if normalized_number in self.registered_numbers:
                already_registered.append(normalized_number)
            else:
                numbers.append(normalized_number)

        if not numbers and not already_registered:
            QMessageBox.warning(self, '警告', '未找到有效号码！')
            return

        # 显示已注册号码的提示并拒绝添加
        if already_registered:
            skip_msg = "以下号码已在注册列表中，不能添加：\n" + "\n".join(already_registered)
            QMessageBox.warning(self, '警告', skip_msg)
            if not numbers:  # 如果所有号码都是重复的
                self.phone_input.clear()
                return

        # 检查是否有号码已在待检测列表中
        duplicates = []
        new_numbers = []
        for number in numbers:
            if number in self.phone_numbers:
                duplicates.append(number)
            else:
                new_numbers.append(number)

        # 显示重复号码的提示
        if duplicates:
            dup_msg = "以下号码已在待检测列表中：\n" + "\n".join(duplicates)
            QMessageBox.warning(self, '警告', dup_msg)

        if not new_numbers:
            self.phone_input.clear()
            return

        # 添加新号码
        self.phone_numbers.extend(new_numbers)
        self.update_numbers_display()
        self.phone_input.clear()

        QMessageBox.information(self, '成功', f'成功添加 {len(new_numbers)} 个新号码！')

    def clear_phone_numbers(self):
        """清空待检测号码列表"""
        try:
            if not self.phone_numbers:
                QMessageBox.information(self, '提示', '检测列表已经是空的！')
                return

            count = len(self.phone_numbers)
            reply = QMessageBox.question(self, '确认', f'确定要清空所有 {count} 个待检测号码吗？',
                                         QMessageBox.Yes | QMessageBox.No)

            if reply == QMessageBox.Yes:
                # 清空列表
                self.phone_numbers = []
                # 更新显示
                self.update_numbers_display()
                # 记录日志
                self.log(f"已清空 {count} 个待检测号码")
                # 通知用户
                QMessageBox.information(self, '成功', f'已清空 {count} 个号码！')
        except Exception as e:
            QMessageBox.warning(self, '错误', f'清空号码时发生错误: {str(e)}')

    def update_numbers_display(self):
        """更新号码显示，使用优化的方法显示大量号码"""
        if isinstance(self.numbers_list, NumbersListWidget):
            # 如果已经使用了优化的显示组件
            self.numbers_list.set_numbers(self.phone_numbers)
        else:
            # 显示前1000个，避免界面卡顿
            self.numbers_list.clear()
            display_limit = 1000
            total = len(self.phone_numbers)

            if total > 0:
                # 显示有限数量的号码
                for i in range(min(display_limit, total)):
                    self.numbers_list.append(self.phone_numbers[i])

                if total > display_limit:
                    self.numbers_list.append(
                        f"\n...\n(仅显示前 {display_limit} 个，共 {total} 个号码)")
                else:
                    self.numbers_list.append(f"\n\n共 {total} 个号码")
            else:
                self.numbers_list.append("没有号码")

    def import_numbers(self):
        file_name, _ = QFileDialog.getOpenFileName(self, '选择号码文件',
                                                   '', '文本文件 (*.txt)')
        if not file_name:
            return

        # 显示进度对话框
        progress_dialog = QDialog(self)
        progress_dialog.setWindowTitle("导入号码")
        progress_dialog.setFixedSize(400, 200)
        progress_layout = QVBoxLayout()

        # 进度显示
        progress_label = QLabel("正在处理文件...")
        progress_bar = QProgressBar()
        progress_bar.setRange(0, 100)

        # 取消按钮
        cancel_btn = QPushButton("取消")

        progress_layout.addWidget(progress_label)
        progress_layout.addWidget(progress_bar)
        progress_layout.addWidget(cancel_btn, 0, Qt.AlignCenter)

        progress_dialog.setLayout(progress_layout)

        # 创建导入线程
        self.import_thread = PhoneNumberImportThread(
            file_path=file_name,
            existing_numbers=self.phone_numbers,
            registered_numbers=self.registered_numbers
        )

        # 连接信号
        self.import_thread.progress_signal.connect(
            lambda current, total: progress_bar.setValue(
                int(current / total * 100))
        )
        self.import_thread.log_signal.connect(
            lambda msg: progress_label.setText(msg)
        )
        self.import_thread.result_signal.connect(
            lambda numbers: self.finish_import(numbers, progress_dialog)
        )

        # 取消按钮
        cancel_btn.clicked.connect(self.cancel_import)

        # 启动线程
        self.import_thread.start()

        # 显示进度对话框
        progress_dialog.exec_()

    def cancel_import(self):
        """取消导入操作"""
        if hasattr(self, 'import_thread') and self.import_thread.isRunning():
            self.import_thread.stop()
            self.log("正在取消导入...")

    def finish_import(self, numbers, dialog):
        """完成导入后的处理"""
        if numbers:
            # 添加到号码列表
            self.phone_numbers.extend(numbers)
            # 更新显示
            self.update_numbers_display()
            # 关闭进度对话框
            dialog.accept()

            QMessageBox.information(self, '成功', f'成功导入 {len(numbers)} 个新号码！')
        else:
            dialog.accept()
            QMessageBox.warning(self, '提示', '没有导入任何新号码！')

    def log(self, message):
        """添加日志消息到状态文本框"""
        current_time = QDateTime.currentDateTime().toString('yyyy-MM-dd hh:mm:ss')
        formatted_message = f"[{current_time}] {message}"
        self.status_text.append(formatted_message)
        # 滚动到最新内容
        self.status_text.verticalScrollBar().setValue(
            self.status_text.verticalScrollBar().maximum()
        )
        # 确保立即更新显示
        QApplication.processEvents()

    def start_check(self):
        try:
            if not self.sessions:
                QMessageBox.warning(self, '警告', '请先添加Session！')
                return

            if not self.phone_numbers:
                QMessageBox.warning(self, '警告', '请先添加要检查的号码！')
                return

            dialog = SendMessageDialog(self)
            if dialog.exec_() == QDialog.Accepted:
                message = dialog.get_message()
                # 清空状态文本框
                self.status_text.clear()
                self.log("开始检测...")

                # 更新UI状态
                self.start_btn.setEnabled(False)
                self.check_activity_btn.setEnabled(False)
                self.stop_btn.setEnabled(True)

                # 创建并启动检查线程
                self.check_thread = CheckThread(
                    self.phone_numbers.copy(), message, self.sessions, self)
                self.check_thread.log_signal.connect(self.log)
                self.check_thread.progress_signal.connect(self.update_progress)
                self.check_thread.session_status_signal.connect(
                    self.update_session_status)
                self.check_thread.check_complete_signal.connect(
                    self.check_completed)
                self.check_thread.start()
        except Exception as e:
            import traceback
            error_msg = f"启动筛选检测时出错: {str(e)}\n{traceback.format_exc()}"
            self.log(f"【严重错误】{error_msg}")
            QMessageBox.critical(
                self, '错误', f"启动检测时发生错误：\n{str(e)}\n\n请检查程序日志获取详细信息。")

            # 恢复UI状态
            self.start_btn.setEnabled(True)
            self.check_activity_btn.setEnabled(True)
            self.stop_btn.setEnabled(False)

    def start_activity_check(self):
        """开始检测用户活跃度"""
        try:
            if not self.sessions:
                QMessageBox.warning(self, '警告', '请先添加Session！')
                return

            if not self.phone_numbers:
                QMessageBox.warning(self, '警告', '请先添加要检查的号码！')
                return

            # 清空状态文本框
            self.status_text.clear()
            self.log("开始检测Telegram用户活跃度...")

            # 更新UI状态
            self.start_btn.setEnabled(False)
            self.check_activity_btn.setEnabled(False)
            self.stop_btn.setEnabled(True)

            # 创建并启动活跃度检测线程
            self.activity_thread = ActivityCheckThread(
                self.phone_numbers.copy(), self.sessions, self)
            self.activity_thread.log_signal.connect(self.log)
            self.activity_thread.progress_signal.connect(self.update_progress)
            self.activity_thread.session_status_signal.connect(
                self.update_session_status)
            self.activity_thread.check_complete_signal.connect(
                self.activity_check_completed)
            self.activity_thread.start()
        except Exception as e:
            import traceback
            error_msg = f"启动活跃度检测时出错: {str(e)}\n{traceback.format_exc()}"
            self.log(f"【严重错误】{error_msg}")
            QMessageBox.critical(
                self, '错误', f"启动活跃度检测时发生错误：\n{str(e)}\n\n请检查程序日志获取详细信息。")

            # 恢复UI状态
            self.start_btn.setEnabled(True)
            self.check_activity_btn.setEnabled(True)
            self.stop_btn.setEnabled(False)

    def activity_check_completed(self, activity_results):
        """活跃度检测完成的处理"""
        self.start_btn.setEnabled(True)
        self.check_activity_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)

        if activity_results:
            self.log('\n【活跃度检测完成】')
            active_count = sum(1 for a in activity_results if a.is_active)
            total_count = len(activity_results)
            active_percentage = (active_count / total_count *
                                 100) if total_count > 0 else 0

            self.log(
                f'【检测结果】总计 {total_count} 个号码, 其中 {active_count} 个活跃 ({active_percentage:.1f}%)')

            # 自动导出为Excel文件
            try:
                # 获取当前时间作为文件名一部分
                current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
                default_filename = f'telegram_activity_{current_time}.xlsx'

                # 保存对话框
                file_name, _ = QFileDialog.getSaveFileName(
                    self,
                    '保存活跃度检测结果',
                    default_filename,
                    'Excel文件 (*.xlsx)'
                )

                if file_name:
                    # 创建DataFrame
                    data = []
                    for user in activity_results:
                        last_seen_str = ""
                        if user.last_seen:
                            last_seen_str = user.last_seen.strftime(
                                "%d日 %H:%M:%S")

                        check_time_str = ""
                        if user.check_time:
                            check_time_str = user.check_time.strftime(
                                "%d日 %H:%M:%S")

                        # 计算时间差
                        time_diff_str = "未知"
                        if user.last_seen and user.check_time:
                            time_diff = user.check_time - user.last_seen
                            days = time_diff.days
                            hours, remainder = divmod(time_diff.seconds, 3600)
                            minutes, seconds = divmod(remainder, 60)

                            # 格式化时间差字符串
                            if days > 0:
                                time_diff_str = f"{days}天{hours}小时{minutes}分"
                            elif hours > 0:
                                time_diff_str = f"{hours}小时{minutes}分"
                            else:
                                time_diff_str = f"{minutes}分{seconds}秒"

                        data.append({
                            '手机号': user.phone_number,
                            '活跃状态': user.activity_status,
                            '最后在线': last_seen_str,
                            '检测时间': check_time_str,
                            '距离最后在线': time_diff_str
                        })

                    # 创建DataFrame并导出
                    df = pd.DataFrame(data)

                    # 设置Excel写入选项
                    writer = pd.ExcelWriter(file_name, engine='openpyxl')
                    df.to_excel(writer, index=False, sheet_name='活跃度检测结果')

                    # 设置固定列宽
                    worksheet = writer.sheets['活跃度检测结果']
                    column_widths = {
                        'A': 25,  # 手机号
                        'B': 15,  # 活跃状态
                        'C': 20,  # 最后在线
                        'D': 20,  # 检测时间
                        'E': 20   # 距离最后在线
                    }

                    # 应用列宽设置
                    for col, width in column_widths.items():
                        worksheet.column_dimensions[col].width = width

                    # 设置标题行样式
                    from openpyxl.styles import Font, Alignment, PatternFill
                    header_font = Font(bold=True, size=11)
                    header_fill = PatternFill(
                        start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                    header_alignment = Alignment(
                        horizontal='center', vertical='center')

                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment

                    # 设置数据行居中对齐
                    for row in worksheet.iter_rows(min_row=2):
                        for cell in row:
                            cell.alignment = Alignment(
                                horizontal='center', vertical='center')

                    # 自动调整行高
                    for row in worksheet.rows:
                        worksheet.row_dimensions[row[0].row].height = 20

                    # 保存文件
                    writer.close()

                    self.log(f'【导出成功】检测结果已导出到Excel文件: {file_name}')
                    QMessageBox.information(
                        self, '导出成功', f'检测结果已成功导出到:\n{file_name}')

                    # 询问是否打开文件
                    open_reply = QMessageBox.question(
                        self,
                        '打开文件',
                        '是否立即打开导出的Excel文件?',
                        QMessageBox.Yes | QMessageBox.No
                    )

                    if open_reply == QMessageBox.Yes:
                        if sys.platform == 'win32':
                            os.startfile(file_name)
                        elif sys.platform == 'darwin':  # macOS
                            os.system(f'open "{file_name}"')
                        else:  # Linux
                            os.system(f'xdg-open "{file_name}"')
            except Exception as e:
                error_msg = f"导出Excel文件时出错: {str(e)}"
                self.log(f"【导出失败】{error_msg}")
                QMessageBox.warning(self, '导出错误', error_msg)
        else:
            self.log('\n【活跃度检测完成】未获取到有效结果')
            QMessageBox.information(self, '完成', '活跃度检测未找到有效结果！')

    def save_active_users(self, activity_results):
        """保存活跃用户到文件"""
        try:
            active_users = [a for a in activity_results if a.is_active]
            if not active_users:
                QMessageBox.information(self, '提示', '没有检测到活跃用户！')
                return

            file_name, _ = QFileDialog.getSaveFileName(
                self,
                '保存活跃用户信息',
                'telegram_active_users.txt',
                '文本文件 (*.txt)'
            )

            if file_name:
                with open(file_name, 'w', encoding='utf-8') as f:
                    f.write(
                        f"# Telegram活跃用户列表 - 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"# 共检测到 {len(active_users)} 个活跃用户\n\n")

                    for user in active_users:
                        # 基本信息：手机号和活跃状态
                        f.write(f"{user.phone_number}")
                        if user.username:
                            f.write(f" @{user.username}")
                        f.write(f" ({user.activity_status})\n")

                self.log(f"已将 {len(active_users)} 个活跃用户信息保存到: {file_name}")
                QMessageBox.information(
                    self, '成功', f'已将 {len(active_users)} 个活跃用户信息保存到:\n{file_name}')
        except Exception as e:
            error_msg = f"保存活跃用户信息时出错: {str(e)}"
            self.log(error_msg)
            QMessageBox.warning(self, '错误', error_msg)

    def stop_check(self):
        """停止检测进程"""
        try:
            # 尝试停止注册检测线程
            if hasattr(self, 'check_thread') and self.check_thread is not None:
                if self.check_thread.isRunning():
                    self.log("正在停止注册检测...")
                    try:
                        self.check_thread.stop()
                    except Exception as e:
                        self.log(f"【停止错误】停止注册检测线程时发生错误: {str(e)}")

            # 尝试停止活跃度检测线程
            if hasattr(self, 'activity_thread') and self.activity_thread is not None:
                if self.activity_thread.isRunning():
                    self.log("正在停止活跃度检测...")
                    try:
                        self.activity_thread.stop()
                    except Exception as e:
                        self.log(f"【停止错误】停止活跃度检测线程时发生错误: {str(e)}")

            # 更新UI状态
            self.stop_btn.setEnabled(False)
            self.start_btn.setEnabled(True)
            self.check_activity_btn.setEnabled(True)
        except Exception as e:
            import traceback
            error_msg = f"停止检测时出错: {str(e)}\n{traceback.format_exc()}"
            self.log(f"【严重错误】{error_msg}")

            # 确保UI恢复正常状态
            try:
                self.stop_btn.setEnabled(False)
                self.start_btn.setEnabled(True)
                self.check_activity_btn.setEnabled(True)
            except:
                pass

            QMessageBox.critical(
                self, '错误', f"停止检测时发生错误：\n{str(e)}\n\n程序将尝试恢复。")

    def closeEvent(self, event):
        """程序关闭前保存设置"""
        self.save_settings()
        super().closeEvent(event)

    def toggle_select_all_sessions(self):
        """全选或取消全选所有会话"""
        # 检查当前是否所有会话都已被选中
        if not self.selected_sessions:
            return  # 如果没有会话，直接返回

        # 计算当前已选中会话的数量
        checked_count = sum(
            1 for checkbox in self.selected_sessions.values() if checkbox.isChecked())
        all_checked = checked_count == len(self.selected_sessions)

        # 如果全部已选中，则取消全选；否则全选
        for checkbox in self.selected_sessions.values():
            checkbox.setChecked(not all_checked)

        # 更新视图
        self.session_table.viewport().update()

        # 显示消息
        if all_checked:
            self.log(f"已取消选择所有会话")
        else:
            self.log(f"已选择所有会话({len(self.selected_sessions)}个)")

    def delete_invalid_sessions(self):
        """删除所有错误或未授权的会话"""
        if not self.sessions:
            QMessageBox.warning(self, '警告', '没有可删除的Session！')
            return

        # 先筛选出需要删除的会话
        invalid_sessions = [
            session for session in self.sessions if session.status == "错误" or session.status == "未授权"]

        if not invalid_sessions:
            QMessageBox.information(self, '提示', '没有发现错误或未授权的会话！')
            return

        reply = QMessageBox.question(
            self,
            '确认',
            f'确定要删除全部 {len(invalid_sessions)} 个错误或未授权的会话吗？\n这将同时删除对应的.session文件！',
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            deleted_count = 0
            error_count = 0
            error_sessions = []

            # 创建会话副本，避免删除时修改列表导致问题
            sessions_to_remove = []

            for session in invalid_sessions:
                try:
                    # 检查文件是否存在及是否可写
                    if os.path.exists(session.file_path):
                        # 检查是否有删除权限
                        if not os.access(session.file_path, os.W_OK):
                            error_msg = f"没有删除权限: {session.file_path}"
                            self.log(error_msg)
                            error_sessions.append(f"{session.name}: 没有文件删除权限")
                            error_count += 1
                            continue

                        try:
                            # 尝试关闭可能打开的文件句柄
                            os.chmod(session.file_path, 0o777)  # 尝试改变文件权限
                            os.remove(session.file_path)
                            self.log(f"已删除文件: {session.file_path}")
                            sessions_to_remove.append(session)
                            deleted_count += 1
                        except PermissionError:
                            self.log(f"权限错误，无法删除: {session.file_path}")
                            error_sessions.append(f"{session.name}: 权限被拒绝")
                            error_count += 1
                        except Exception as e:
                            self.log(f"删除文件时出错: {str(e)}")
                            error_sessions.append(f"{session.name}: {str(e)}")
                            error_count += 1
                    else:
                        # 文件不存在，只从列表中删除
                        self.log(f"文件不存在，仅从列表移除: {session.file_path}")
                        sessions_to_remove.append(session)
                        deleted_count += 1

                except Exception as e:
                    self.log(f"处理Session {session.name} 时出错: {str(e)}")
                    error_sessions.append(f"{session.name}: {str(e)}")
                    error_count += 1

            # 从会话列表中移除已删除的会话
            for session in sessions_to_remove:
                if session in self.sessions:
                    self.sessions.remove(session)

            # 更新界面显示
            self.update_session_list()
            self.save_settings()  # 保存设置

            # 显示结果消息
            if deleted_count > 0:
                if error_count > 0:
                    error_details = "\n".join(error_sessions)
                    QMessageBox.warning(
                        self,
                        '部分完成',
                        f'已成功删除 {deleted_count} 个不可用会话！\n但有 {error_count} 个会话删除失败:\n\n{error_details}\n\n请确保程序有权限删除文件，并且文件未被其他程序占用。'
                    )
                else:
                    QMessageBox.information(
                        self,
                        '成功',
                        f'已成功删除所有 {deleted_count} 个不可用会话！'
                    )
            else:
                QMessageBox.critical(
                    self,
                    '失败',
                    f'删除操作全部失败，原因如下:\n\n{os.linesep.join(error_sessions)}\n\n请尝试以管理员身份运行程序，或手动删除文件。'
                )

    def update_session_status(self, session_path, status, error=None):
        """更新会话状态"""
        for session in self.sessions:
            if session.file_path == session_path:
                session.status = status
                if status == "正在运行":
                    session.last_used = time.time()
                elif status == "空闲":
                    session.total_checks += 1
                break

        # 使用延迟更新机制，避免频繁更新造成界面卡顿
        if not hasattr(self, '_update_ui_timer'):
            self._update_ui_timer = QTimer()
            self._update_ui_timer.timeout.connect(self._delayed_update_ui)
            self._update_ui_timer.setSingleShot(True)

        # 如果计时器未运行，则启动它
        if not self._update_ui_timer.isActive():
            self._update_ui_timer.start(500)  # 500毫秒后更新UI

    def _delayed_update_ui(self):
        """延迟执行UI更新，合并多个更新请求"""
        # 立即更新显示
        self.update_session_list()
        # 确保UI更新
        QApplication.processEvents()

    def update_progress(self, current, total):
        """更新进度信息"""
        progress = (current / total) * 100 if total > 0 else 0

        # 只在整数百分比变化时输出日志，减少过多的输出
        if int(progress) % 5 == 0 or current == total or current == 1:
            self.log(f"【筛号进度】已检测: {current}/{total} ({progress:.1f}%)")

        # 使用动画更新进度条
        if hasattr(self, 'progress_bar') and isinstance(self.progress_bar, StyledProgressBar):
            self.progress_bar.updateValue(progress)
        else:
            self.progress_bar.setValue(int(progress))

    def check_completed(self, registered_numbers):
        """检查完成的处理"""
        try:
            self.start_btn.setEnabled(True)
            self.check_activity_btn.setEnabled(True)
            self.stop_btn.setEnabled(False)

            if registered_numbers:
                self.log('\n【检测完成】发现已注册号码:')
                for number in registered_numbers:
                    normalized_number = self.normalize_phone_number(number)
                    self.log(normalized_number)
                    self.registered_numbers.add(normalized_number)

                # 保存结果到文件
                try:
                    # 获取phone_register.txt文件路径 - 优先使用可执行文件所在目录
                    if getattr(sys, 'frozen', False):
                        # 如果是打包后的exe运行
                        base_dir = os.path.dirname(sys.executable)
                        phone_register_path = os.path.join(
                            base_dir, 'phone_register.txt')
                    else:
                        # 如果是开发环境运行
                        phone_register_path = 'phone_register.txt'

                    with open(phone_register_path, 'a', encoding='utf-8') as f:
                        for number in registered_numbers:
                            normalized_number = self.normalize_phone_number(
                                number)
                            f.write(f"{normalized_number}\n")
                    self.log(f'\n【保存完成】已注册号码保存至: {phone_register_path}')

                    # 询问是否打开结果文件
                    reply = QMessageBox.question(
                        self,
                        '检测完成',
                        f'检测完成! 共发现 {len(registered_numbers)} 个已注册号码\n是否打开结果文件查看？',
                        QMessageBox.Yes | QMessageBox.No
                    )

                    if reply == QMessageBox.Yes:
                        # 使用系统默认程序打开文件
                        try:
                            if sys.platform == 'win32':
                                os.startfile(phone_register_path)
                            elif sys.platform == 'darwin':  # macOS
                                os.system(f'open "{phone_register_path}"')
                            else:  # Linux
                                os.system(f'xdg-open "{phone_register_path}"')
                        except Exception as open_error:
                            self.log(f'\n【打开失败】无法打开结果文件: {str(open_error)}')
                            QMessageBox.warning(
                                self, '错误', f'无法打开结果文件: {str(open_error)}')

                except Exception as save_error:
                    self.log(f'\n【保存失败】{str(save_error)}')
                    QMessageBox.warning(
                        self, '错误', f'保存结果文件失败: {str(save_error)}')
            else:
                self.log('\n【检测完成】本次检查未发现已注册号码')
                QMessageBox.information(self, '完成', '检测完成，未发现已注册号码！')
        except Exception as e:
            import traceback
            error_msg = f"处理检测结果时出错: {str(e)}\n{traceback.format_exc()}"
            self.log(f"【严重错误】{error_msg}")
            QMessageBox.critical(
                self, '错误', f"处理检测结果时发生错误：\n{str(e)}\n\n但您的检测数据不会丢失。")

    def open_result_file_location(self):
        """打开结果文件所在位置"""
        # 获取phone_register.txt文件路径 - 优先使用可执行文件所在目录
        if getattr(sys, 'frozen', False):
            # 如果是打包后的exe运行
            base_dir = os.path.dirname(sys.executable)
            file_path = os.path.join(base_dir, 'phone_register.txt')
        else:
            # 如果是开发环境运行
            file_path = os.path.abspath('phone_register.txt')

        if os.path.exists(file_path):
            # 根据不同操作系统打开文件所在位置
            if sys.platform == 'win32':
                os.system(f'explorer /select,"{file_path}"')
            elif sys.platform == 'darwin':  # macOS
                os.system(f'open -R "{file_path}"')
            else:  # Linux
                os.system(f'xdg-open "{os.path.dirname(file_path)}"')
        else:
            QMessageBox.warning(self, '提示', f'结果文件不存在！\n路径：{file_path}')

    def show_config_dialog(self):
        """显示配置对话框"""
        dialog = ConfigDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            # 重新加载配置 - 获取配置文件路径
            if getattr(sys, 'frozen', False):
                # 如果是打包后的exe运行
                base_dir = os.path.dirname(sys.executable)
                config_path = os.path.join(base_dir, 'config.ini')
            else:
                # 如果是开发环境运行
                config_path = 'config.ini'

            self.config.read(config_path, encoding='utf-8')
            QMessageBox.information(self, '提示', '配置已更新，部分设置可能需要重启程序后生效。')

    def get_api_credentials(self):
        """获取API凭证"""
        return (
            self.config.getint('API', 'api_id', fallback=2040),
            self.config.get('API', 'api_hash',
                            fallback='b18441a1ff607e10a989891a5462e627')
        )

    def load_settings(self):
        """从JSON文件加载设置"""
        try:
            # 获取设置文件路径 - 优先使用可执行文件所在目录
            if getattr(sys, 'frozen', False):
                # 如果是打包后的exe运行
                base_dir = os.path.dirname(sys.executable)
                settings_path = os.path.join(base_dir, self.settings_file)
            else:
                # 如果是开发环境运行
                settings_path = self.settings_file

            if os.path.exists(settings_path):
                with open(settings_path, 'r', encoding='utf-8') as f:
                    settings = json.load(f)

                # 加载Session设置
                if 'sessions' in settings:
                    for session_data in settings['sessions']:
                        try:
                            # 加载时使用保存的值，不使用配置文件的默认值
                            session = SessionStatus.from_dict(session_data)
                            if os.path.exists(session.file_path):
                                self.sessions.append(session)
                        except Exception as e:
                            print(f"加载Session设置出错: {str(e)}")
        except Exception as e:
            print(f"加载设置文件出错: {str(e)}")

    def save_settings(self):
        """保存设置到JSON文件"""
        try:
            settings = {
                'sessions': [session.to_dict() for session in self.sessions]
            }

            # 获取设置文件路径 - 优先使用可执行文件所在目录
            if getattr(sys, 'frozen', False):
                # 如果是打包后的exe运行
                base_dir = os.path.dirname(sys.executable)
                settings_path = os.path.join(base_dir, self.settings_file)
            else:
                # 如果是开发环境运行
                settings_path = self.settings_file

            with open(settings_path, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"保存设置出错: {str(e)}")

    def closeEvent(self, event):
        """程序关闭前保存设置"""
        self.save_settings()
        super().closeEvent(event)


if __name__ == '__main__':
    # 添加全局异常处理器
    import traceback
    import sys

    # 原始excepthook
    sys._excepthook = sys.excepthook

    # 自定义异常处理函数
    def exception_hook(exctype, value, tb):
        # 将异常信息格式化为字符串
        error_msg = ''.join(traceback.format_exception(exctype, value, tb))
        print(f"发生未捕获异常:\n{error_msg}")

        # 如果GUI已经初始化，显示错误对话框
        try:
            from PyQt5.QtWidgets import QMessageBox
            if 'app' in globals() and 'gui' in globals() and gui is not None:
                QMessageBox.critical(gui, "程序错误",
                                     f"程序发生错误，但已被拦截，不会闪退:\n\n{str(value)}\n\n详细信息已记录到程序日志")
                # 尝试记录到GUI日志
                if hasattr(gui, 'log'):
                    gui.log(f"【严重错误】程序发生未捕获异常: {str(value)}")
        except Exception as dialog_error:
            print(f"显示错误对话框时出错: {str(dialog_error)}")

        # 调用原始异常处理
        sys._excepthook(exctype, value, tb)

    # 设置全局异常处理器
    sys.excepthook = exception_hook

    # 全局声明应用和主窗口对象
    app = None
    gui = None

    try:
        app = QApplication(sys.argv)
        app.setStyle('Fusion')  # 使用Fusion风格获得更现代的外观

        # 创建启动画面
        from PyQt5.QtGui import QPixmap, QPainter, QColor, QFont, QPen
        from PyQt5.QtCore import Qt, QTimer

        # 创建自定义启动画面
        class CustomSplashScreen(QWidget):
            def __init__(self):
                super().__init__()
                self.setWindowFlags(Qt.FramelessWindowHint |
                                    Qt.WindowStaysOnTopHint)
                self.setAttribute(Qt.WA_TranslucentBackground)
                self.setFixedSize(400, 300)
                self.setStyleSheet("background-color: transparent;")
                self.progress = 0
                self.timer = QTimer(self)
                self.timer.timeout.connect(self.update_progress)
                self.timer.start(30)  # 每30毫秒更新一次
                self.loading_texts = [
                    "初始化界面...",
                    "加载配置文件...",
                    "准备SESSION会话...",
                    "检查系统资源...",
                    "初始化完成!"
                ]
                self.current_text_index = 0
                self.text_timer = QTimer(self)
                self.text_timer.timeout.connect(self.update_text)
                self.text_timer.start(600)  # 每600毫秒更新一次文本

                # 显示在屏幕中央
                self.center()

            def center(self):
                """将窗口居中"""
                screen = QApplication.desktop().screenGeometry()
                size = self.geometry()
                self.move(int((screen.width() - size.width()) / 2),
                          int((screen.height() - size.height()) / 2))

            def update_progress(self):
                """更新进度"""
                self.progress += 1
                if self.progress > 100:
                    self.timer.stop()
                    return
                self.update()

            def update_text(self):
                """更新加载文本"""
                self.current_text_index = (
                    self.current_text_index + 1) % len(self.loading_texts)
                self.update()

            def paintEvent(self, event):
                """绘制启动画面"""
                painter = QPainter(self)
                painter.setRenderHint(QPainter.Antialiasing)

                # 背景
                painter.setPen(Qt.NoPen)
                painter.setBrush(QColor(30, 30, 30, 200))
                painter.drawRoundedRect(
                    0, 0, self.width(), self.height(), 15, 15)

                # 标题
                painter.setPen(QPen(QColor("#2196F3"), 1))
                title_font = QFont("Arial", 16, QFont.Bold)
                painter.setFont(title_font)
                painter.drawText(QRect(0, 40, self.width(), 40),
                                 Qt.AlignCenter, "Telegram消息工具")

                # 副标题/版本
                painter.setPen(QPen(QColor("#BBDEFB"), 1))
                subtitle_font = QFont("Arial", 10)
                painter.setFont(subtitle_font)
                painter.drawText(QRect(0, 80, self.width(), 20),
                                 Qt.AlignCenter, "版本 1.0.1")

                # 加载文本
                text_font = QFont("Arial", 11)
                painter.setFont(text_font)
                painter.setPen(QPen(QColor("#E3F2FD"), 1))
                painter.drawText(QRect(0, 150, self.width(), 30), Qt.AlignCenter,
                                 self.loading_texts[self.current_text_index])

                # 进度条背景
                painter.setPen(Qt.NoPen)
                painter.setBrush(QColor(60, 60, 60, 150))
                painter.drawRoundedRect(50, 200, 300, 15, 7, 7)

                # 进度条
                painter.setBrush(QColor("#4CAF50"))
                progress_width = int(300 * (self.progress / 100))
                painter.drawRoundedRect(50, 200, progress_width, 15, 7, 7)

                # 版权信息
                copyright_font = QFont("Arial", 8)
                painter.setFont(copyright_font)
                painter.setPen(QPen(QColor("#78909C"), 1))
                painter.drawText(QRect(0, 250, self.width(), 20),
                                 Qt.AlignCenter, "© 2025 zat 版权所有")

        # 创建启动屏幕
        splash = CustomSplashScreen()
        splash.show()
        app.processEvents()  # 确保启动画面显示

        # 使用定时器延迟加载主窗口
        def show_main_window():
            global gui
            try:
                gui = TelegramGUI()
                gui.show()
                if splash and splash.isVisible():
                    splash.hide()
            except Exception as e:
                print(f"创建主窗口时出错: {str(e)}\n{traceback.format_exc()}")
                if splash and splash.isVisible():
                    splash.hide()
                QMessageBox.critical(None, "启动错误",
                                     f"启动程序时出错:\n\n{str(e)}")
                sys.exit(1)

        # 延时1.5秒后显示主窗口
        QTimer.singleShot(1500, show_main_window)

        sys.exit(app.exec_())

    except Exception as startup_error:
        # 如果在启动过程中发生任何错误，显示错误消息框
        print(f"启动错误: {str(startup_error)}\n{traceback.format_exc()}")
        try:
            if app:
                QMessageBox.critical(None, "启动错误",
                                     f"程序启动失败:\n\n{str(startup_error)}")
        except:
            pass

        sys.exit(1)
